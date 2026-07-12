"""
Enrich M1–M6 Excel files using local xlsb bank feed files.
No DB changes. Reads xlsb files as lookup tables and patches missing columns.

Sources:
  DEMAND xlsb  → EMI Start/End, Installments, DPD(Mar), Bucket(Mar), AM, Last Coll Date, NPA Amt
  POS xlsb     → DPD(Feb), Bucket(Feb)

Column positions in 01_RAW_DATA (0-indexed):
  4  = Area Manager
  19 = EMI Start Date
  20 = EMI End Date
  21 = Installment Count
  25 = DPD
  26 = Bank Bucket
  31 = Last Month Paid Date
  33 = NPA Amt
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import pyxlsb
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
from shutil import copy2
import time

# ── File paths ────────────────────────────────────────────────────────────────
TEMPLATE    = 'd:/LOS/M1.xlsx'
DEMAND_XLSB = 'd:/LOS/data/ATYATI-FEDERAL-DEMAND-A3-MAR-2026_29-03-2026.xlsb'
POS_XLSB    = 'd:/LOS/data/FEDERAL-POS-MONTH-END-FEB-26.xlsb'

M_FILES = [
    ('d:/LOS/M1.xlsx', '2025-10'),
    ('d:/LOS/M2.xlsx', '2025-11'),
    ('d:/LOS/M3.xlsx', '2025-12'),
    ('d:/LOS/M4.xlsx', '2026-01'),
    ('d:/LOS/M5.xlsx', '2026-02'),
    ('d:/LOS/M6.xlsx', '2026-03'),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def xl_date(v):
    """Convert Excel serial date float to YYYY-MM-DD string."""
    if v is None or v == '':
        return ''
    try:
        d = datetime(1899, 12, 30) + timedelta(days=int(float(v)))
        return d.strftime('%Y-%m-%d')
    except Exception:
        return str(v) if v else ''


def acct_str(v):
    """Normalise account number to string (xlsb stores as float)."""
    if v is None:
        return None
    try:
        return str(int(float(v)))
    except Exception:
        return str(v).strip()


def safe(v):
    """Return value or empty string."""
    if v is None:
        return ''
    return str(v).strip() if isinstance(v, str) else v


# ── Load DEMAND xlsb ──────────────────────────────────────────────────────────

def load_demand(path):
    """
    Returns dict keyed by account_no (str):
      {acct: (emi_start, emi_end, period, dpd, bank_bucket, arrear_amt, am, last_coll_date)}
    """
    print(f"Loading demand file: {path}", flush=True)
    t0 = time.time()
    lookup = {}
    skipped = 0

    with pyxlsb.open_workbook(path) as wb:
        with wb.get_sheet('Demand') as ws:
            rows = ws.rows()
            next(rows)  # skip header
            for row in rows:
                vals = [c.v for c in row]
                if len(vals) < 44:
                    skipped += 1
                    continue
                acct = acct_str(vals[2])
                if not acct:
                    continue

                npa_flag   = str(vals[19]).strip().upper() if vals[19] else ''
                arrear_amt = vals[40] if npa_flag == 'Y' else None

                lookup[acct] = (
                    xl_date(vals[13]),   # EMI Start Date
                    xl_date(vals[14]),   # EMI End Date
                    safe(vals[15]),      # PERIOD (In Months)
                    safe(vals[16]),      # DPD
                    safe(vals[17]),      # Bank Bucket
                    arrear_amt,          # NPA Amt (only when NPA=Y)
                    safe(vals[43]),      # AM
                    xl_date(vals[46]) if len(vals) > 46 else '',  # Last coll date
                )

    print(f"  Loaded {len(lookup):,} accounts in {time.time()-t0:.1f}s (skipped {skipped})", flush=True)
    return lookup


# ── Load POS xlsb ─────────────────────────────────────────────────────────────

def load_pos(path):
    """
    Returns dict keyed by account_no (str):
      {acct: (dpd_days, bucket)}
    """
    print(f"Loading POS file: {path}", flush=True)
    t0 = time.time()
    lookup = {}

    with pyxlsb.open_workbook(path) as wb:
        with wb.get_sheet('Sheet1') as ws:
            rows = ws.rows()
            next(rows)  # skip header
            for row in rows:
                vals = [c.v for c in row]
                if len(vals) < 20:
                    continue
                acct = acct_str(vals[4])
                if not acct:
                    continue
                lookup[acct] = (
                    safe(vals[18]),  # Dpd Days
                    safe(vals[19]),  # Bucket
                )

    print(f"  Loaded {len(lookup):,} accounts in {time.time()-t0:.1f}s", flush=True)
    return lookup


# ── Enrich one M file ─────────────────────────────────────────────────────────

def enrich_file(filepath, data_month, demand, pos):
    print(f"\n[{filepath}] {data_month}", flush=True)
    t0 = time.time()

    use_demand_dpd  = (data_month == '2026-03')  # Mar 2026 DPD/Bucket from demand
    use_pos_dpd     = (data_month == '2026-02')  # Feb 2026 DPD/Bucket from POS

    wb = load_workbook(filepath)
    ws = wb['01_RAW_DATA']

    total = ws.max_row - 3  # rows 1-3 are headers
    matched_demand = 0
    matched_pos    = 0
    filled = {
        'AM': 0, 'EMI Start': 0, 'EMI End': 0, 'Installments': 0,
        'DPD': 0, 'Bucket': 0, 'Last Coll': 0, 'NPA': 0
    }

    for row_idx in range(4, ws.max_row + 1):
        # Account number is col 12 (1-indexed)
        acct_cell = ws.cell(row_idx, 12)
        acct = str(acct_cell.value).strip() if acct_cell.value else None
        if not acct:
            continue

        # ── Demand file enrichment (all months) ──────────────────────────────
        d = demand.get(acct)
        if d:
            matched_demand += 1
            emi_start, emi_end, period, dpd, bucket, npa_amt, am, last_coll = d

            def fill(col, val, key):
                if val not in ('', None) and ws.cell(row_idx, col).value in ('', None, 0):
                    ws.cell(row_idx, col).value = val
                    filled[key] += 1

            fill(5,  am,        'AM')
            fill(20, emi_start, 'EMI Start')
            fill(21, emi_end,   'EMI End')
            fill(22, period,    'Installments')
            fill(32, last_coll, 'Last Coll')
            if npa_amt not in ('', None):
                fill(34, npa_amt, 'NPA')

            # DPD & Bucket — only for March (demand file is March data)
            if use_demand_dpd:
                fill(26, dpd,    'DPD')
                fill(27, bucket, 'Bucket')

        # ── POS file enrichment (Feb only) ───────────────────────────────────
        if use_pos_dpd:
            p = pos.get(acct)
            if p:
                matched_pos += 1
                dpd_days, bkt = p
                if dpd_days not in ('', None) and ws.cell(row_idx, 26).value in ('', None, 0):
                    ws.cell(row_idx, 26).value = dpd_days
                    filled['DPD'] += 1
                if bkt not in ('', None) and ws.cell(row_idx, 27).value in ('', None, 0):
                    ws.cell(row_idx, 27).value = bkt
                    filled['Bucket'] += 1

        if row_idx % 50000 == 4:
            print(f"  {row_idx-3:,}/{total:,} rows processed...", flush=True)

    print(f"  Matched: demand={matched_demand:,}  pos={matched_pos:,}", flush=True)
    print(f"  Filled: {filled}", flush=True)
    print(f"  Saving...", flush=True)
    wb.save(filepath)
    wb.close()
    print(f"  Done in {time.time()-t0:.1f}s", flush=True)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=== Client JLG — Enrich M1–M6 from local files ===\n", flush=True)
    total_start = time.time()

    demand = load_demand(DEMAND_XLSB)
    pos    = load_pos(POS_XLSB)
    print(flush=True)

    for filepath, data_month in M_FILES:
        enrich_file(filepath, data_month, demand, pos)

    print(f"\n✓ All done in {(time.time()-total_start)/60:.1f} min", flush=True)


if __name__ == '__main__':
    main()
