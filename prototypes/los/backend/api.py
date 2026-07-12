"""
api.py — Flask REST API for the Loan Threat System (LOS) dashboard.
Serves on http://localhost:5000

Endpoints
---------
GET  /api/health
GET  /api/summary
GET  /api/trends
GET  /api/buckets?month=YYYY-MM
GET  /api/bucket-movement
GET  /api/top-critical
GET  /api/flags?month=YYYY-MM
GET  /api/pool/<pool_name>?month=YYYY-MM
GET  /api/hierarchy?month=YYYY-MM
GET  /api/models/metrics
POST /api/predict/<model_name>      body: { features: {...} }
GET  /api/loans                     ?search=&month=&bucket=&page=&size=
GET  /api/loan/<loan_id>?month=
POST /api/export/pool/<pool>?month=  → Excel bytes
"""

import os, io, json, pickle, warnings
from functools import lru_cache
from datetime import datetime

import numpy as np
import pandas as pd
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

warnings.filterwarnings("ignore")

import sys
sys.path.insert(0, os.path.dirname(__file__))
from data_processor import (
    load_dataset, get_feature_columns, get_portfolio_summary,
    get_bucket_distribution, get_monthly_trends, get_top_critical_loans,
    get_bucket_movement, get_flag_table, get_pool_data, get_hierarchy,
)

app  = Flask(__name__)
CORS(app)

MODELS_DIR   = os.path.join(os.path.dirname(__file__), "models")
FEATURE_COLS = get_feature_columns()

# ─────────────────────────────────────────────────────────────────────────────
# Data + model caching
# ─────────────────────────────────────────────────────────────────────────────

_df: pd.DataFrame = None

def get_df() -> pd.DataFrame:
    global _df
    if _df is None:
        _df = load_dataset()
    return _df


@lru_cache(maxsize=8)
def _load_model(name: str) -> dict:
    path = os.path.join(MODELS_DIR, f"{name}.pkl")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Model not found: {name}")
    with open(path, "rb") as f:
        return pickle.load(f)


def _load_metrics() -> dict:
    path = os.path.join(MODELS_DIR, "metrics.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def err(msg: str, code: int = 400):
    return jsonify({"error": msg}), code


def ok(data):
    return jsonify(data)


def _predict_single(model_bundle: dict, feature_dict: dict) -> dict:
    feats = [feature_dict.get(f, 0) for f in model_bundle["features"]]
    X     = np.array(feats).reshape(1, -1)
    m     = model_bundle["model"]

    if model_bundle["type"] == "classifier":
        pred  = m.predict(X)[0]
        proba = m.predict_proba(X)[0]
        if "label_encoder" in model_bundle:
            le   = model_bundle["label_encoder"]
            cls  = le.inverse_transform([pred])[0]
            return {"prediction": cls,
                    "confidence": round(float(proba.max()), 4),
                    "probabilities": {
                        c: round(float(p), 4)
                        for c, p in zip(le.classes_, proba)
                    }}
        else:
            classes = model_bundle.get("classes", [str(i) for i in range(len(proba))])
            return {"prediction": classes[pred],
                    "confidence": round(float(proba.max()), 4),
                    "probabilities": {
                        c: round(float(p), 4) for c, p in zip(classes, proba)
                    }}
    else:
        val = float(m.predict(X)[0])
        rng = model_bundle.get("output_range", [None, None])
        if rng[0] is not None:
            val = float(np.clip(val, rng[0], rng[1]))
        return {"prediction": round(val, 4)}


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/health")
def health():
    return ok({"status": "ok", "timestamp": datetime.utcnow().isoformat()})


@app.route("/api/summary")
def summary():
    df = get_df()
    return ok(get_portfolio_summary(df))


@app.route("/api/trends")
def trends():
    df = get_df()
    return ok(get_monthly_trends(df))


@app.route("/api/buckets")
def buckets():
    df    = get_df()
    month = request.args.get("month")
    return ok(get_bucket_distribution(df, month))


@app.route("/api/bucket-movement")
def bucket_movement():
    df = get_df()
    return ok(get_bucket_movement(df))


@app.route("/api/top-critical")
def top_critical():
    df = get_df()
    n  = int(request.args.get("n", 10))
    return ok(get_top_critical_loans(df, n))


@app.route("/api/flags")
def flags():
    df    = get_df()
    month = request.args.get("month")
    return ok(get_flag_table(df, month))


@app.route("/api/pool/<pool_name>")
def pool(pool_name):
    df    = get_df()
    month = request.args.get("month")
    valid = ["REGULAR", "MODERATE", "RISK", "CRITICAL", "VERY CRITICAL"]
    name  = pool_name.upper().replace("-", " ")
    if name not in valid:
        return err(f"Unknown pool: {pool_name}. Valid: {valid}")
    return ok(get_pool_data(df, name, month))


@app.route("/api/hierarchy")
def hierarchy():
    df    = get_df()
    month = request.args.get("month")
    return ok(get_hierarchy(df, month))


@app.route("/api/models/metrics")
def model_metrics():
    return ok(_load_metrics())


@app.route("/api/loans")
def loans():
    df     = get_df()
    month  = request.args.get("month")
    bucket = request.args.get("bucket")
    search = request.args.get("search", "").strip().lower()
    page   = int(request.args.get("page", 1))
    size   = int(request.args.get("size", 50))

    subset = df.copy()
    if month:
        subset = subset[subset["month_label"] == month]
    else:
        subset = subset[subset["month_label"] == subset["month_label"].max()]
    if bucket:
        subset = subset[subset["bucket"] == bucket.upper()]
    if search:
        mask = (
            subset["loan_id"].str.lower().str.contains(search, na=False) |
            subset["customer_id"].str.lower().str.contains(search, na=False) |
            subset["branch"].str.lower().str.contains(search, na=False)
        )
        subset = subset[mask]

    total  = len(subset)
    subset = subset.iloc[(page - 1) * size: page * size]
    cols   = ["loan_id", "customer_id", "branch", "rm", "state",
              "product", "dpd", "bucket", "alert_level",
              "outstanding", "demand", "collected", "shortfall",
              "collection_eff", "structural_risk", "cb_score"]
    return ok({
        "total": total, "page": page, "size": size,
        "pages": (total + size - 1) // size,
        "loans": subset[cols].to_dict(orient="records"),
    })


@app.route("/api/loan/<loan_id>")
def loan_detail(loan_id):
    df    = get_df()
    month = request.args.get("month")
    subset = df if not month else df[df["month_label"] == month]
    row   = subset[subset["loan_id"] == loan_id]
    if row.empty:
        return err(f"Loan {loan_id} not found", 404)
    # All months for this loan
    history = df[df["loan_id"] == loan_id].sort_values("month_label")
    return ok({
        "current": row.iloc[0].to_dict(),
        "history": history[["month_label", "dpd", "bucket", "alert_level",
                              "collected", "demand", "collection_eff"]
                            ].to_dict(orient="records")
    })


# ─────────────────────────────────────────────────────────────────────────────
# Prediction endpoints
# ─────────────────────────────────────────────────────────────────────────────

MODEL_NAMES = {
    "default-predictor":  "default_predictor",
    "risk-scorer":        "risk_scorer",
    "bucket-forecaster":  "bucket_forecaster",
    "alert-engine":       "alert_engine",
    "collection-engine":  "collection_engine",
}

@app.route("/api/predict/<model_name>", methods=["POST"])
def predict(model_name):
    key = MODEL_NAMES.get(model_name)
    if not key:
        return err(f"Unknown model: {model_name}. Valid: {list(MODEL_NAMES)}")
    try:
        bundle  = _load_model(key)
    except FileNotFoundError as e:
        return err(str(e), 404)

    body = request.get_json(silent=True) or {}
    feats = body.get("features", {})
    if not feats:
        # If loan_id provided, fetch features from dataset
        loan_id = body.get("loan_id")
        month   = body.get("month")
        if loan_id:
            df = get_df()
            sub = df if not month else df[df["month_label"] == month]
            row = sub[sub["loan_id"] == loan_id]
            if row.empty:
                return err(f"Loan {loan_id} not found", 404)
            feats = row.iloc[0][FEATURE_COLS].fillna(0).to_dict()
        else:
            return err("Provide 'features' dict or 'loan_id' in request body")

    try:
        result  = _predict_single(bundle, feats)
        fi      = bundle["model"].feature_importances_ if hasattr(bundle["model"], "feature_importances_") else []
        fi_dict = {}
        if len(fi):
            pairs = sorted(zip(bundle["features"], fi), key=lambda x: -x[1])
            fi_dict = {k: round(float(v), 4) for k, v in pairs[:10]}

        return ok({
            "model":             model_name,
            "loan_id":           body.get("loan_id"),
            **result,
            "feature_importance": fi_dict,
        })
    except Exception as e:
        return err(f"Prediction failed: {e}", 500)


# ─────────────────────────────────────────────────────────────────────────────
# Export endpoint
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/export/pool/<pool_name>")
def export_pool(pool_name):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return err("openpyxl not installed", 500)

    df    = get_df()
    month = request.args.get("month")
    name  = pool_name.upper().replace("-", " ")
    data  = get_pool_data(df, name, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{name} Pool"

    fill_map = {
        "REGULAR":      "C6EFCE", "MODERATE":      "FFEB9C",
        "RISK":         "FFCC99", "CRITICAL":       "FFC7CE",
        "VERY CRITICAL":"CC0000",
    }
    hdr_fill = PatternFill("solid", fgColor=fill_map.get(name, "CCCCCC"))
    hdr_font = Font(bold=True, color="000000" if name != "VERY CRITICAL" else "FFFFFF")

    headers = ["Loan ID", "Customer ID", "Branch", "DPD",
               "Outstanding", "Shortfall", "Coll. Eff %", "Bucket"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(data["loans"], 2):
        ws.cell(ri, 1, row.get("loan_id"))
        ws.cell(ri, 2, row.get("customer_id"))
        ws.cell(ri, 3, row.get("branch"))
        ws.cell(ri, 4, row.get("dpd"))
        ws.cell(ri, 5, row.get("outstanding"))
        ws.cell(ri, 6, row.get("shortfall"))
        ws.cell(ri, 7, row.get("collection_eff"))
        ws.cell(ri, 8, row.get("bucket"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{name.replace(' ', '_')}_pool_{month or 'latest'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  LOS Flask API  —  http://localhost:5001")
    print("=" * 55)
    print("[API] Loading dataset…")
    get_df()            # warm the cache
    print("[API] Dataset ready. Starting server…\n")
    app.run(host="0.0.0.0", port=5001, debug=False)
