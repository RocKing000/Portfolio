"""
data_processor.py
Reads the 6 LOS Excel files (M1-M6) and synthesises a unified loan-level
dataset that can be consumed by train_models.py and the Flask API.

Because many fields in the Excel files are ⚠ BLANK (no DB source yet),
this module generates statistically realistic synthetic records so that
the 5 ML models can be trained end-to-end for the demo.
All synthetic logic is isolated in _generate_synthetic_portfolio() and
can be replaced by real DB queries later.
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..")   # D:/LOS/
MONTHS = [
    {"file": "M1", "label": "2025-10", "display": "October 2025"},
    {"file": "M2", "label": "2025-11", "display": "November 2025"},
    {"file": "M3", "label": "2025-12", "display": "December 2025"},
    {"file": "M4", "label": "2026-01", "display": "January 2026"},
    {"file": "M5", "label": "2026-02", "display": "February 2026"},
    {"file": "M6", "label": "2026-03", "display": "March 2026"},
]

NPA_THRESHOLD = 120
BUCKET_BINS   = [0, 0, 30, 60, 90, 120, np.inf]
BUCKET_LABELS = ["B0", "B1", "B2", "B3", "B4+", "NPA"]
ALERT_MAP     = {"B0": "REGULAR", "B1": "MODERATE", "B2": "RISK",
                 "B3": "CRITICAL", "B4+": "VERY CRITICAL", "NPA": "VERY CRITICAL"}

# ─────────────────────────────────────────────────────────────────────────────
# Synthetic portfolio generator
# ─────────────────────────────────────────────────────────────────────────────

def _generate_synthetic_portfolio(n_loans: int = 2000, seed: int = 42) -> pd.DataFrame:
    """
    Returns a DataFrame with one row per loan × month combination (12 000 rows
    for 2 000 loans × 6 months) containing all features needed by the models.
    """
    rng = np.random.default_rng(seed)

    loan_ids       = [f"LN{str(i).zfill(6)}" for i in range(1, n_loans + 1)]
    customer_ids   = [f"CU{str(i).zfill(6)}" for i in range(1, n_loans + 1)]

    states         = ["Maharashtra", "Karnataka", "Telangana", "Tamil Nadu",
                      "Gujarat", "Rajasthan", "Uttar Pradesh", "West Bengal"]
    branches       = [f"BR{str(i).zfill(3)}" for i in range(1, 51)]
    rms            = [f"RM{str(i).zfill(3)}" for i in range(1, 201)]
    products       = ["JLG-MICRO", "JLG-AGRI", "JLG-SME", "JLG-LIVELIHOOD"]

    # Loan-level static attributes
    loan_amount    = rng.integers(5_000, 100_000, n_loans).astype(float)
    emi_amount     = (loan_amount / rng.integers(12, 48, n_loans)).round(2)
    tenure_months  = rng.integers(12, 48, n_loans)
    age            = rng.integers(21, 65, n_loans)
    group_size     = rng.integers(5, 20, n_loans)
    income_monthly = rng.integers(5_000, 40_000, n_loans).astype(float)
    cb_score       = rng.integers(300, 850, n_loans)
    state          = rng.choice(states, n_loans)
    branch         = rng.choice(branches, n_loans)
    rm             = rng.choice(rms, n_loans)
    product        = rng.choice(products, n_loans)
    gender         = rng.choice(["F", "M"], n_loans, p=[0.85, 0.15])

    # Base DPD for month 0 — skewed toward low DPD (healthy portfolio)
    base_dpd = rng.exponential(scale=18, size=n_loans).clip(0, 365).astype(int)

    rows = []
    for mi, month_info in enumerate(MONTHS):
        # DPD drifts slightly each month
        dpd = (base_dpd + rng.integers(-5, 15, n_loans)).clip(0, 365)

        outstanding = (loan_amount * rng.uniform(0.2, 1.0, n_loans)).round(2)
        demand      = emi_amount.copy()
        collected   = (demand * rng.uniform(0.5, 1.05, n_loans)).clip(0, demand * 1.2).round(2)
        shortfall   = (demand - collected).clip(0)
        coll_eff    = (collected / demand * 100).round(2)

        # Bucket from DPD
        bucket_idx  = np.searchsorted([0, 1, 31, 61, 91, 121], dpd, side="right") - 1
        bucket_idx  = bucket_idx.clip(0, len(BUCKET_LABELS) - 1)
        bucket      = np.array(BUCKET_LABELS)[bucket_idx]

        # Alert level
        alert       = np.array([ALERT_MAP[b] for b in bucket])

        # Structural risk score (0-100, derived features)
        struct_risk = (
            dpd * 0.4
            + (shortfall / demand.clip(1) * 100) * 0.3
            + (1 - cb_score / 850) * 30
        ).clip(0, 100).round(2)

        # Triggered flags
        flg_death       = (rng.random(n_loans) < 0.005).astype(int)
        flg_unreachable = (rng.random(n_loans) < 0.04).astype(int)
        flg_geo         = (rng.random(n_loans) < 0.02).astype(int)
        flg_new_loan    = (rng.random(n_loans) < 0.08).astype(int)
        flg_preclosure  = (rng.random(n_loans) < 0.03).astype(int)
        flg_family_loan = (rng.random(n_loans) < 0.06).astype(int)

        # Default label (target for Default Predictor model)
        default_label = (dpd >= 90).astype(int)

        for i in range(n_loans):
            rows.append({
                "month_label":      month_info["label"],
                "month_display":    month_info["display"],
                "loan_id":          loan_ids[i],
                "customer_id":      customer_ids[i],
                "state":            state[i],
                "branch":           branch[i],
                "rm":               rm[i],
                "product":          product[i],
                "gender":           gender[i],
                "age":              int(age[i]),
                "group_size":       int(group_size[i]),
                "income_monthly":   float(income_monthly[i]),
                "cb_score":         int(cb_score[i]),
                "loan_amount":      float(loan_amount[i]),
                "emi_amount":       float(emi_amount[i]),
                "tenure_months":    int(tenure_months[i]),
                "outstanding":      float(outstanding[i]),
                "dpd":              int(dpd[i]),
                "demand":           float(demand[i]),
                "collected":        float(collected[i]),
                "shortfall":        float(shortfall[i]),
                "collection_eff":   float(coll_eff[i]),
                "bucket":           bucket[i],
                "alert_level":      alert[i],
                "structural_risk":  float(struct_risk[i]),
                "flg_death":        int(flg_death[i]),
                "flg_unreachable":  int(flg_unreachable[i]),
                "flg_geo":          int(flg_geo[i]),
                "flg_new_loan":     int(flg_new_loan[i]),
                "flg_preclosure":   int(flg_preclosure[i]),
                "flg_family_loan":  int(flg_family_loan[i]),
                "default_label":    int(default_label[i]),
            })

    df = pd.DataFrame(rows)
    print(f"[DataProcessor] Synthetic portfolio: {len(df):,} rows "
          f"({n_loans} loans × {len(MONTHS)} months)")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Excel reader (picks up any real values that exist in the LOS files)
# ─────────────────────────────────────────────────────────────────────────────

def read_excel_files() -> pd.DataFrame:
    """
    Attempts to extract numerical values from the 01_RAW_DATA sheet of each
    workbook. Falls back gracefully to empty if cells contain only headers/notes.
    """
    records = []
    for m in MONTHS:
        path = os.path.join(EXCEL_DIR, f"{m['file']}.xlsx")
        if not os.path.exists(path):
            print(f"[DataProcessor] ⚠  {path} not found — skipping")
            continue
        try:
            wb  = openpyxl.load_workbook(path, data_only=True)
            ws  = wb["01_RAW_DATA"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[3]]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if all(v is None for v in row):
                    continue
                rec = dict(zip(headers, row))
                rec["month_label"]   = m["label"]
                rec["month_display"] = m["display"]
                records.append(rec)
        except Exception as exc:
            print(f"[DataProcessor] ⚠  Could not parse {m['file']}.xlsx: {exc}")
    return pd.DataFrame(records) if records else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def load_dataset() -> pd.DataFrame:
    """
    Returns the best available dataset:
    - Real rows from Excel files (if present) merged with synthetic data,
    - Otherwise pure synthetic data.
    """
    excel_df = read_excel_files()
    synth_df = _generate_synthetic_portfolio()

    if not excel_df.empty:
        print(f"[DataProcessor] Merged {len(excel_df)} real rows with synthetic data")
        df = pd.concat([excel_df, synth_df], ignore_index=True)
    else:
        print("[DataProcessor] No real rows found — using synthetic portfolio only")
        df = synth_df

    # Drop rows that have no usable label data (real Excel stub rows)
    df = df.dropna(subset=["bucket", "alert_level", "default_label"]).reset_index(drop=True)

    # Encode categoricals for ML
    df["bucket_code"]     = pd.Categorical(df["bucket"],
                                           categories=BUCKET_LABELS).codes
    _alert_cats = list(dict.fromkeys(ALERT_MAP.values()))   # unique, order-preserved
    df["alert_code"]      = pd.Categorical(df["alert_level"],
                                           categories=_alert_cats).codes
    df["product_code"]    = pd.Categorical(df["product"]).codes
    df["state_code"]      = pd.Categorical(df["state"]).codes
    return df


def get_feature_columns() -> list:
    return [
        "age", "group_size", "income_monthly", "cb_score",
        "loan_amount", "emi_amount", "tenure_months", "outstanding",
        "dpd", "demand", "collected", "shortfall", "collection_eff",
        "flg_death", "flg_unreachable", "flg_geo",
        "flg_new_loan", "flg_preclosure", "flg_family_loan",
        "product_code", "state_code",
    ]


def get_portfolio_summary(df: pd.DataFrame) -> dict:
    latest = df[df["month_label"] == df["month_label"].max()]
    prev   = df[df["month_label"] == sorted(df["month_label"].unique())[-2]] \
             if df["month_label"].nunique() > 1 else latest

    critical_count   = int((latest["alert_level"].isin(["CRITICAL", "VERY CRITICAL"])).sum())
    prev_critical    = int((prev["alert_level"].isin(["CRITICAL", "VERY CRITICAL"])).sum())
    exposure_at_risk = float(latest.loc[latest["alert_level"].isin(["CRITICAL", "VERY CRITICAL"]),
                                        "outstanding"].sum())
    prev_exposure    = float(prev.loc[prev["alert_level"].isin(["CRITICAL", "VERY CRITICAL"]),
                                      "outstanding"].sum())
    coll_eff         = float(latest["collection_eff"].mean())
    prev_coll_eff    = float(prev["collection_eff"].mean())
    npa_rate         = float((latest["bucket"] == "NPA").sum() / max(len(latest), 1) * 100)
    prev_npa_rate    = float((prev["bucket"] == "NPA").sum() / max(len(prev), 1) * 100)

    return {
        "critical_count":      critical_count,
        "critical_trend":      critical_count - prev_critical,
        "exposure_at_risk":    round(exposure_at_risk, 2),
        "exposure_trend":      round(exposure_at_risk - prev_exposure, 2),
        "collection_eff":      round(coll_eff, 2),
        "collection_eff_trend":round(coll_eff - prev_coll_eff, 2),
        "npa_rate":            round(npa_rate, 2),
        "npa_rate_trend":      round(npa_rate - prev_npa_rate, 2),
    }


def get_bucket_distribution(df: pd.DataFrame, month: str = None) -> dict:
    subset = df[df["month_label"] == month] if month else df
    counts = subset["bucket"].value_counts().to_dict()
    return {b: counts.get(b, 0) for b in BUCKET_LABELS}


def get_monthly_trends(df: pd.DataFrame) -> list:
    rows = []
    for m in sorted(df["month_label"].unique()):
        sub = df[df["month_label"] == m]
        rows.append({
            "month":          m,
            "demand":         round(sub["demand"].sum(), 2),
            "collected":      round(sub["collected"].sum(), 2),
            "shortfall":      round(sub["shortfall"].sum(), 2),
            "collection_eff": round(sub["collection_eff"].mean(), 2),
            "npa_count":      int((sub["bucket"] == "NPA").sum()),
            "total_loans":    len(sub),
        })
    return rows


def get_top_critical_loans(df: pd.DataFrame, n: int = 10) -> list:
    latest = df[df["month_label"] == df["month_label"].max()]
    top    = latest.nlargest(n, "dpd")[
        ["loan_id", "customer_id", "branch", "dpd",
         "outstanding", "bucket", "alert_level", "collection_eff"]
    ]
    return top.to_dict(orient="records")


def get_bucket_movement(df: pd.DataFrame) -> dict:
    months = sorted(df["month_label"].unique())
    if len(months) < 2:
        return {}
    prev_m = months[-2]
    curr_m = months[-1]
    prev   = df[df["month_label"] == prev_m][["loan_id", "bucket"]].rename(
                 columns={"bucket": "from_bucket"})
    curr   = df[df["month_label"] == curr_m][["loan_id", "bucket"]].rename(
                 columns={"bucket": "to_bucket"})
    merged = prev.merge(curr, on="loan_id")
    matrix = {}
    for fb in BUCKET_LABELS:
        matrix[fb] = {}
        for tb in BUCKET_LABELS:
            matrix[fb][tb] = int(
                ((merged["from_bucket"] == fb) & (merged["to_bucket"] == tb)).sum()
            )
    return matrix


def get_flag_table(df: pd.DataFrame, month: str = None) -> list:
    subset = df[df["month_label"] == month] if month else \
             df[df["month_label"] == df["month_label"].max()]
    cols = ["loan_id", "customer_id", "branch", "rm", "dpd", "outstanding",
            "flg_death", "flg_unreachable", "flg_geo",
            "flg_new_loan", "flg_preclosure", "flg_family_loan", "alert_level"]
    return subset[cols].to_dict(orient="records")


def get_pool_data(df: pd.DataFrame, pool: str, month: str = None) -> dict:
    subset = df[df["month_label"] == month] if month else \
             df[df["month_label"] == df["month_label"].max()]
    pool_df = subset[subset["alert_level"] == pool.upper()]
    return {
        "count":       len(pool_df),
        "outstanding": round(pool_df["outstanding"].sum(), 2),
        "shortfall":   round(pool_df["shortfall"].sum(), 2),
        "avg_dpd":     round(pool_df["dpd"].mean(), 2) if len(pool_df) else 0,
        "loans":       pool_df[["loan_id", "customer_id", "branch",
                                 "dpd", "outstanding", "shortfall",
                                 "collection_eff", "bucket"]
                               ].head(100).to_dict(orient="records"),
    }


def get_hierarchy(df: pd.DataFrame, month: str = None) -> dict:
    subset = df[df["month_label"] == month] if month else \
             df[df["month_label"] == df["month_label"].max()]
    result = {}
    for state, sg in subset.groupby("state"):
        result[state] = {
            "total": len(sg),
            "npa":   int((sg["bucket"] == "NPA").sum()),
            "coll_eff": round(sg["collection_eff"].mean(), 2),
            "branches": {}
        }
        for branch, bg in sg.groupby("branch"):
            result[state]["branches"][branch] = {
                "total":    len(bg),
                "npa":      int((bg["bucket"] == "NPA").sum()),
                "coll_eff": round(bg["collection_eff"].mean(), 2),
            }
    return result


if __name__ == "__main__":
    df = load_dataset()
    print(f"\n[DataProcessor] Dataset shape : {df.shape}")
    print(f"[DataProcessor] Columns       : {list(df.columns)}")
    print(f"\n[DataProcessor] Portfolio summary (latest month):")
    import json
    print(json.dumps(get_portfolio_summary(df), indent=2))
    print(f"\n[DataProcessor] Bucket distribution (latest):")
    print(json.dumps(get_bucket_distribution(df), indent=2))
