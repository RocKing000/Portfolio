"""
distribute_data.py
Reads the real LTS data from M1.xlsx (01_RAW_DATA sheet),
samples 5,000 rows, and writes them into each of the 6 monthly
workbooks (M1–M6) — preserving all other sheets exactly as-is.

The Data Month (YYYY-MM) column is updated per file so every
workbook carries its own month label.

Run AFTER closing M1.xlsx in Excel:
    python D:/LOS/distribute_data.py
"""

import os, copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

LOS_DIR = r"D:\LOS"
SAMPLE_SIZE = 5000
MONTH_COL_HEADER = "Data Month (YYYY-MM)"

MONTHS = [
    {"file": "M1", "label": "2025-10", "display": "October 2025"},
    {"file": "M2", "label": "2025-11", "display": "November 2025"},
    {"file": "M3", "label": "2025-12", "display": "December 2025"},
    {"file": "M4", "label": "2026-01", "display": "January 2026"},
    {"file": "M5", "label": "2026-02", "display": "February 2026"},
    {"file": "M6", "label": "2026-03", "display": "March 2026"},
]


# ── Step 1: Read 5000 rows from M1's RAW_DATA ──────────────────────────────

def read_source(path: str) -> dict:
    print(f"[1/2] Reading source: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["01_RAW_DATA"]

    # Rows 1-3: title / group-headers / column-headers (keep as-is)
    meta_rows = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        meta_rows.append(list(row))

    headers = meta_rows[2]   # row 3 = column names

    # Find month column index
    month_col_idx = None
    for i, h in enumerate(headers):
        if h and MONTH_COL_HEADER in str(h):
            month_col_idx = i
            break

    if month_col_idx is None:
        raise ValueError(f"Could not find '{MONTH_COL_HEADER}' in row 3")

    # Read data rows (row 4 onwards), skip note/sample stub rows from template
    data_rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = list(row)
        # Skip blank rows and the template note row
        if vals[0] is None:
            continue
        if str(vals[0]).startswith("⚠") or str(vals[0]).startswith("NOTE"):
            continue
        data_rows.append(vals)
        if len(data_rows) >= SAMPLE_SIZE:
            break

    wb.close()
    print(f"   Sampled {len(data_rows):,} rows  |  {len(headers)} columns  |  month_col={month_col_idx}")
    return {
        "meta_rows":    meta_rows,
        "headers":      headers,
        "data_rows":    data_rows,
        "month_col_idx": month_col_idx,
        "n_cols":       len(headers),
    }


# ── Step 2: Write into each monthly workbook ───────────────────────────────

def write_month(source: dict, month_info: dict):
    path = os.path.join(LOS_DIR, f"{month_info['file']}.xlsx")
    print(f"\n[2/2] Writing {month_info['file']}.xlsx  ({month_info['display']}) …")

    wb = openpyxl.load_workbook(path)          # full load (write mode)

    # Drop and recreate the RAW_DATA sheet to start clean
    if "01_RAW_DATA" in wb.sheetnames:
        idx = wb.sheetnames.index("01_RAW_DATA")
        del wb["01_RAW_DATA"]
    else:
        idx = 0
    ws = wb.create_sheet("01_RAW_DATA", idx)
    ws.sheet_view.showGridLines = False

    n_cols = source["n_cols"]

    # ── Row 1: Title (merged, dark navy) ───────────────────────────────────
    title_text = f"01 — RAW DATA  |  {month_info['display']}  [{month_info['label']}]"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.fill      = PatternFill("solid", fgColor="1F3864")
    cell.font      = Font(bold=True, color="FFFFFF", size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Row 2: Group headers (copy from source row 2) ──────────────────────
    group_colors = ["2E4A7A", "1F6B3A", "800000", "5C4A72", "7B3F00", "1F6B3A"]
    seg_size = max(1, n_cols // len(group_colors))
    for col_idx in range(1, n_cols + 1):
        grp = min((col_idx - 1) // seg_size, len(group_colors) - 1)
        src_val = source["meta_rows"][1][col_idx - 1] if col_idx - 1 < len(source["meta_rows"][1]) else None
        c = ws.cell(row=2, column=col_idx, value=src_val)
        c.fill      = PatternFill("solid", fgColor=group_colors[grp])
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ───────────────────────────────────────────────
    for col_idx, h in enumerate(source["headers"], 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.fill      = PatternFill("solid", fgColor="D9E1F2")
        c.font      = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = Border(
            bottom=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="DDDDDD"),
        )
    ws.row_dimensions[3].height = 30

    # ── Rows 4+: Data rows (with month label swapped) ──────────────────────
    month_col_idx = source["month_col_idx"]
    thin = Border(right=Side(style="thin", color="EEEEEE"),
                  bottom=Side(style="thin", color="F5F5F5"))

    for r_offset, data_row in enumerate(source["data_rows"]):
        row_num = 4 + r_offset
        for col_idx, val in enumerate(data_row, 1):
            # Overwrite the month column with this file's label
            write_val = month_info["label"] if (col_idx - 1) == month_col_idx else val
            c = ws.cell(row=row_num, column=col_idx, value=write_val)
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal="left")
            c.border    = thin
        if r_offset % 1000 == 0:
            print(f"   {r_offset:,} / {len(source['data_rows']):,} rows written…", end="\r")

    print(f"   {len(source['data_rows']):,} / {len(source['data_rows']):,} rows written.    ")

    # ── Column widths ───────────────────────────────────────────────────────
    col_widths = [18, 18, 16, 20, 16, 16, 8, 20, 22, 20,
                  14, 20, 12, 22, 14, 14, 14, 14, 14, 12,
                  12, 10, 14, 14, 14, 8, 12, 12, 18, 14,
                  14, 16, 18, 14, 14]
    for i, w in enumerate(col_widths[:n_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=4, column=1)

    wb.save(path)
    size_mb = os.path.getsize(path) / 1024 / 1024
    print(f"   Saved  →  {path}  ({size_mb:.1f} MB)")


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    m1_path = os.path.join(LOS_DIR, "M1.xlsx")

    if not os.path.exists(m1_path):
        raise FileNotFoundError(f"M1.xlsx not found at {m1_path}")

    # Check if file is locked by Excel
    lock = os.path.join(LOS_DIR, "~$M1.xlsx")
    if os.path.exists(lock):
        print("⚠  M1.xlsx is still open in Excel. Please close it first, then re-run.")
        raise SystemExit(1)

    source = read_source(m1_path)

    print(f"\n[Writing] 5,000 rows × 6 files …\n{'─'*50}")
    for m in MONTHS:
        write_month(source, m)

    print(f"\n{'─'*50}")
    print("Done. All 6 files updated:")
    for m in MONTHS:
        p = os.path.join(LOS_DIR, f"{m['file']}.xlsx")
        sz = os.path.getsize(p) / 1024
        print(f"  {m['file']}.xlsx  →  {sz:.0f} KB  ({m['display']})")
